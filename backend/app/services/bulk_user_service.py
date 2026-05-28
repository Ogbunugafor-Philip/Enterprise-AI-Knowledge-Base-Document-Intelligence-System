import io
import re
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.department import Department
from app.models.user import User
from app.schemas.user_management import BulkUserError, BulkUserUploadResponse

REQUIRED_COLUMNS = {"first_name", "last_name", "email", "department_name", "role_name"}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def parse_excel_file(file_bytes: bytes) -> tuple[list[dict], list[BulkUserError]]:
    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
    except Exception as exc:
        return [], [BulkUserError(row_number=0, email="", error_reason=f"Failed to open Excel file: {exc}")]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [BulkUserError(row_number=0, email="", error_reason="Excel file is empty")]

    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    missing = REQUIRED_COLUMNS - set(header)
    if missing:
        return [], [BulkUserError(row_number=1, email="", error_reason=f"Missing required columns: {', '.join(sorted(missing))}")]

    col_index = {name: header.index(name) for name in REQUIRED_COLUMNS}
    parsed: list[dict] = []
    errors: list[BulkUserError] = []

    for row_num, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        try:
            record = {col: (str(row[idx]).strip() if row[idx] is not None else "") for col, idx in col_index.items()}
            record["_row"] = row_num
            parsed.append(record)
        except Exception as exc:
            errors.append(BulkUserError(row_number=row_num, email="", error_reason=str(exc)))

    return parsed, errors


def validate_bulk_user_row(
    row: dict,
    seen_emails: set[str],
) -> tuple[bool, str]:
    first_name = row.get("first_name", "").strip()
    last_name = row.get("last_name", "").strip()
    email = row.get("email", "").strip().lower()
    role_name = row.get("role_name", "").strip()

    if not first_name:
        return False, "first_name is required"
    if not last_name:
        return False, "last_name is required"
    if not email:
        return False, "email is required"
    if not EMAIL_RE.match(email):
        return False, f"Invalid email format: {email}"
    if email in seen_emails:
        return False, f"Duplicate email in file: {email}"
    if not role_name:
        return False, "role_name is required"

    return True, ""


async def process_bulk_upload(
    db: AsyncSession,
    organization_id: UUID,
    file_bytes: bytes,
    created_by_user_id: UUID | None,
) -> BulkUserUploadResponse:
    from app.models.role import Role
    from app.core.permissions import RoleEnum, normalize_role
    from app.services.user_management_service import create_user
    from app.models.audit import AuditLog

    parsed_rows, parse_errors = parse_excel_file(file_bytes)

    dept_cache: dict[str, UUID | None] = {}
    role_cache: dict[str, UUID | None] = {}
    seen_emails: set[str] = set()

    created_users = []
    row_errors = list(parse_errors)

    for row in parsed_rows:
        row_num = row["_row"]
        email = row.get("email", "").strip().lower()

        is_valid, error_msg = validate_bulk_user_row(row, seen_emails)
        if not is_valid:
            row_errors.append(BulkUserError(row_number=row_num, email=email, error_reason=error_msg))
            continue

        seen_emails.add(email)

        existing = await db.execute(
            select(User).where(User.organization_id == organization_id, User.email == email)
        )
        if existing.scalar_one_or_none() is not None:
            row_errors.append(BulkUserError(row_number=row_num, email=email, error_reason="Email already registered in organization"))
            continue

        dept_name = row.get("department_name", "").strip()
        if dept_name not in dept_cache:
            dept_result = await db.execute(
                select(Department).where(
                    Department.organization_id == organization_id,
                    Department.name == dept_name,
                )
            )
            dept_obj = dept_result.scalar_one_or_none()
            dept_cache[dept_name] = dept_obj.id if dept_obj else None

        dept_id = dept_cache.get(dept_name)
        if dept_name and dept_id is None:
            row_errors.append(BulkUserError(row_number=row_num, email=email, error_reason=f"Department not found: {dept_name}"))
            continue

        role_name = row.get("role_name", "").strip()
        if role_name not in role_cache:
            role_result = await db.execute(
                select(Role).where(Role.name == role_name, Role.organization_id == organization_id)
            )
            role_obj = role_result.scalar_one_or_none()
            role_cache[role_name] = role_obj.id if role_obj else None

        role_id = role_cache.get(role_name)
        if role_name and role_id is None:
            row_errors.append(BulkUserError(row_number=row_num, email=email, error_reason=f"Role not found: {role_name}"))
            continue

        try:
            user = await create_user(
                db=db,
                first_name=row["first_name"],
                last_name=row["last_name"],
                email=email,
                organization_id=organization_id,
                role_id=role_id,
                department_id=dept_id,
                send_welcome_email=True,
                created_by_user_id=created_by_user_id,
            )
            from app.schemas.user_management import UserCreateResponse
            created_users.append(
                UserCreateResponse(
                    id=user.id,
                    first_name=user.first_name,
                    last_name=user.last_name,
                    email=user.email,
                    department_id=user.department_id,
                    organization_id=user.organization_id,
                    role=user.role.name if user.role else None,
                    is_active=user.is_active,
                    is_verified=user.is_verified,
                    created_at=user.created_at,
                )
            )
        except Exception as exc:
            row_errors.append(BulkUserError(row_number=row_num, email=email, error_reason=str(exc)))

    total_rows = len(parsed_rows)
    db.add(
        AuditLog(
            organization_id=organization_id,
            user_id=created_by_user_id,
            action="BULK_USER_UPLOAD",
            resource_type="user",
            new_value={
                "total_rows": total_rows,
                "created": len(created_users),
                "failed": len(row_errors),
            },
        )
    )
    await db.flush()

    return BulkUserUploadResponse(
        total_rows=total_rows,
        successfully_created=len(created_users),
        failed_rows=len(row_errors),
        errors=row_errors,
        created_users=created_users,
    )


def generate_bulk_upload_template() -> bytes:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font, PatternFill  # type: ignore

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        headers = ["first_name", "last_name", "email", "department_name", "role_name"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

        example = ["Jane", "Doe", "jane.doe@example.com", "Engineering", "USER"]
        for col_idx, value in enumerate(example, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    except Exception:
        return b""
